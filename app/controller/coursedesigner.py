import io
import json

from flask import Blueprint, render_template, request, session
from openpyxl import load_workbook

from app.controller.common import search_inner, visualize_inner
from app.models.base import db
from app.models.coursedesigner import Coursedesigner
from flask_login import LoginManager, login_required, login_user, logout_user, UserMixin

from app.models.course import Course, Cilo, Assessment
from datetime import datetime

log_file = open('course.log', 'a', encoding='utf-8')

coursedesignerBP = Blueprint('coursedesigner', __name__)


@coursedesignerBP.route('coursedesigner_home')
@login_required
def coursedesigner_home():
    user_id = session['_user_id']
    cs = Coursedesigner.query.filter(Coursedesigner.ID_num == user_id).first()
    name = cs.Name
    prog = cs.programme
    return render_template('index_c.html', name=name, type="course degisner")


@coursedesignerBP.route('index')
def index():
    user_id = session['_user_id']
    name = Coursedesigner.query.filter(Coursedesigner.ID_num == user_id).first().Name
    return render_template('welcome.html', name=name)


@coursedesignerBP.route('search')
def search():
    search_word = request.args.get('search')
    search_type = request.args.get('a')
    return search_inner(search_word, search_type)


@coursedesignerBP.route('/create', methods=['GET', 'POST'])
def create():
    if request.method == 'GET':
        user_id = session['_user_id']
        return render_template('createnewcourse.html')
    else:
        CourseName = request.form.get('CourseName')
        CourseCode = request.form.get('CourseCode')
        Year = request.form.get('Year')
        Programme = request.form.get('Programme')
        Type = request.form.get('Type')
        cilo_1 = request.form.get('CILO1')
        cilo_2 = request.form.get('CILO2')
        cilo_3 = request.form.get('CILO3')
        as_name_list = request.form.getlist('as_name')
        as_weight_list = request.form.getlist('as_weight')
        as_cilos_list = request.form.getlist('as_cilo')
        print(request.form)

        if not CourseName or \
                not CourseCode or \
                not Year or \
                not Programme or \
                not Type:
            return render_template('createnewcourse.html', title="Missing fields of course")

        if not request.files.get('cilo_file') and not cilo_1:
            return render_template('createnewcourse.html', title="At least one CILO is needed")

        if request.files.get('cilo_file') and cilo_1:
            return render_template('createnewcourse.html', title="[CILO] Don't input text and upload file at the same time")

        as_list = []
        for i in range(len(as_name_list)):
            if not as_name_list[i] \
                    or not as_weight_list[i] \
                    or not as_cilos_list[i]:
                # 跳过没有填写的条目
                continue
            as_list.append((as_name_list[i], as_weight_list[i], as_cilos_list[i]))

        if request.files.get('assess_file'):
            if len(as_list) > 0:
                return render_template('createnewcourse.html', title="[Assessment] Don't input text and upload file at the same time")
            for as_dict in parse_excel(request.files.get('assess_file').stream):
                as_list.append((as_dict['Evaluation Method'],
                                as_dict['Percentage'].strip("%"),
                                as_dict['CILOs'],))

        weight_sum = 0
        try:
            for _, _weight, _ in as_list:
                weight_sum += int(_weight)
            if weight_sum != 100:
                return render_template('createnewcourse.html', title="Sum of weights is %d, rather than 100" % (weight_sum))
        except Exception:
            return render_template('createnewcourse.html', title="Sum of weights is not 100")

        # 增加 Course
        course = Course()
        course.name = CourseName
        course.code = CourseCode
        course.academic_year = Year
        course.programme = Programme
        course._type = Type
        db.session.add(course)
        db.session.commit()

        # 增加 CILO 和 Course对应关系
        # 优先读文件
        cilos = []
        if request.files.get('cilo_file'):
            for idx, cilo_dict in enumerate(parse_excel(request.files.get('cilo_file').stream)):
                _ = Cilo(cilo_dict['Upon successful completion of the course, students should be able to:'])
                db.session.add(_)
                db.session.commit()
                cilos.append(_)
                if idx == 0:
                    course.cilo1_id = _.id
                elif idx == 1:
                    course.cilo2_id = _.id
                else:
                    course.cilo3_id = _.id
        else:
            if cilo_1:
                _ = Cilo(cilo_1)
                db.session.add(_)
                db.session.commit()
                cilos.append(_)
                course.cilo1_id = _.id
            if cilo_2:
                _ = Cilo(cilo_2)
                db.session.add(_)
                db.session.commit()
                cilos.append(_)
                course.cilo2_id = _.id
            if cilo_3:
                _ = Cilo(cilo_3)
                db.session.add(_)
                db.session.commit()
                cilos.append(_)
                course.cilo3_id = _.id

        db.session.commit()

        # 添加分数
        for _name, _weight, _cilo in as_list:
            db.session.add(Assessment(course, _name, _weight, _cilo))
        db.session.commit()

        log_file.write('{}: {}\n'.format(datetime.now(), "create a course"))
        log_file.flush()

        return render_template('createnewcourse.html', title="Success!")


@coursedesignerBP.route('all_courses')
def all_courses():
    return render_template('all_courses.html', courses=Course.query.all())


@coursedesignerBP.route('edit_course/<course_id>', methods=['GET', 'POST'])
def edit_course(course_id):
    course_id = int(course_id)
    course = Course.query.get(course_id)
    errmsg = None

    if request.method == 'POST':
        if request.form.get('CILO1_id') is not None:
            data = request.form.get('CILO1_id')
            if len(data) == 0:
                course.cilo1_id = None
            elif Cilo.query.get(course.cilo2_id) and data == Cilo.query.get(course.cilo2_id).name \
                    or Cilo.query.get(course.cilo3_id) and data == Cilo.query.get(course.cilo3_id).name:
                errmsg = "ERROR: duplicated CILO"
            elif Cilo.query.filter(Cilo.name == data).first():
                course.cilo1_id = Cilo.query.filter(Cilo.name == data).first().id
            else:
                _ = Cilo(data)
                db.session.add(_)
                db.session.commit()
                course.cilo1_id = _.id
        elif request.form.get('CILO2_id') is not None:
            data = request.form.get('CILO2_id')
            if len(data) == 0:
                course.cilo2_id = None
            elif Cilo.query.get(course.cilo1_id) and data == Cilo.query.get(course.cilo1_id).name \
                    or Cilo.query.get(course.cilo3_id) and data == Cilo.query.get(course.cilo3_id).name:
                errmsg = "ERROR: duplicated CILO"
            elif Cilo.query.filter(Cilo.name == data).first():
                course.cilo2_id = Cilo.query.filter(Cilo.name == data).first().id
            else:
                _ = Cilo(data)
                db.session.add(_)
                db.session.commit()
                course.cilo2_id = _.id
        elif request.form.get('CILO3_id') is not None:
            data = request.form.get('CILO3_id')
            if len(data) == 0:
                course.cilo3_id = None
            elif Cilo.query.get(course.cilo1_id) and data == Cilo.query.get(course.cilo1_id).name \
                    or Cilo.query.get(course.cilo2_id) and data == Cilo.query.get(course.cilo2_id).name:
                errmsg = "ERROR: duplicated CILO"
            elif Cilo.query.filter(Cilo.name == data).first():
                course.cilo3_id = Cilo.query.filter(Cilo.name == data).first().id
            else:
                _ = Cilo(data)
                db.session.add(_)
                db.session.commit()
                course.cilo3_id = _.id
        elif request.form.get('pre1_id'):
            data = int(request.form.get('pre1_id'))
            if data == 0:
                course.pre_course1_id = None
            elif data == course_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite Course"
            elif data == course.pre_course2_id or data == course.pre_course3_id:
                errmsg = "ERROR: duplicated Prerequisite Course"
            else:
                course.pre_course1_id = data

        elif request.form.get('pre2_id'):
            data = int(request.form.get('pre2_id'))
            if data == 0:
                course.pre_course2_id = None
            elif data == course_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite Course"
            elif data == course.pre_course1_id or data == course.pre_course3_id:
                errmsg = "ERROR: duplicated Prerequisite Course"
            else:
                course.pre_course2_id = data
        elif request.form.get('pre3_id'):
            data = int(request.form.get('pre3_id'))
            if data == 0:
                course.pre_course3_id = None
            elif data == course_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite Course"
            elif data == course.pre_course1_id or data == course.pre_course2_id:
                errmsg = "ERROR: duplicated Prerequisite Course"
            else:
                course.pre_course3_id = data
        elif len(request.form.getlist('as_name')) > 0:
            as_list = []
            as_name_list = request.form.getlist('as_name')
            as_weight_list = request.form.getlist('as_weight')
            as_cilos_list = request.form.getlist('as_cilo')

            for i in range(len(as_name_list)):
                if not as_name_list[i] \
                        or not as_weight_list[i] \
                        or not as_cilos_list[i]:
                    # 跳过没有填写的条目
                    continue
                as_list.append((as_name_list[i], as_weight_list[i], as_cilos_list[i]))
            weight_sum = 0
            try:
                for _, _weight, _ in as_list:
                    weight_sum += int(_weight)
                if weight_sum != 100:
                    errmsg = "ERROR: Sum of weights is %d, rather than 100" % (weight_sum)
                else:
                    for _name, _weight, _cilo in as_list:
                        db.session.add(Assessment(course, _name, _weight, _cilo))
                    db.session.commit()

            except Exception:
                errmsg = "ERROR: Sum of weights is not 100"

        db.session.commit()
        print(request.form)
        if errmsg:
            log_file.write('{}: {}\n'.format(datetime.now(), "course is modified fail, reason: " + errmsg))
        else:
            log_file.write('{}: {}\n'.format(datetime.now(), "course is modified successful"))
        log_file.flush()

    cilo1_str = Cilo.query.get(course.cilo1_id).name if course.cilo1_id else ""
    cilo2_str = Cilo.query.get(course.cilo2_id).name if course.cilo2_id else ""
    cilo3_str = Cilo.query.get(course.cilo3_id).name if course.cilo3_id else ""

    course_all = Course.query.all()
    ass_list = Assessment.latest_before(int(datetime.now().timestamp()), course_id)
    return render_template('edit_course.html',
                           title=errmsg,
                           course=course,
                           course_all=course_all,
                           ass_list=ass_list,
                           cilo1_str=cilo1_str,
                           cilo2_str=cilo2_str,
                           cilo3_str=cilo3_str,
                           )

@coursedesignerBP.route('edit_cilo/<cilo_id>', methods=['GET', 'POST'])
def edit_cilo(cilo_id):
    cilo_id = int(cilo_id)
    cilo = Cilo.query.get(cilo_id)
    errmsg = None
    if request.method == "POST":
        if request.form.get('pre1_id'):
            data = int(request.form.get('pre1_id'))
            if data == 0:
                cilo.pre_cilo1 = None
            elif data == cilo_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite CILO"
            elif data == cilo.pre_cilo2 or data == cilo.pre_cilo3:
                errmsg = "ERROR: duplicated Prerequisite CILO"
            else:
                cilo.pre_cilo1 = data
        elif request.form.get('pre2_id'):
            data = int(request.form.get('pre2_id'))
            if data == 0:
                cilo.pre_cilo2 = None
            elif data == cilo_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite CILO"
            elif data == cilo.pre_cilo1 or data == cilo.pre_cilo3:
                errmsg = "ERROR: duplicated Prerequisite CILO"
            else:
                cilo.pre_cilo2 = data
        elif request.form.get('pre3_id'):
            data = int(request.form.get('pre3_id'))
            if data == 0:
                cilo.pre_cilo3 = None
            elif data == cilo_id:
                errmsg = "ERROR: Cannot assign self as Prerequisite CILO"
            elif data == cilo.pre_cilo1 or data == cilo.pre_cilo2:
                errmsg = "ERROR: duplicated Prerequisite CILO"
            else:
                cilo.pre_cilo3 = data
        db.session.commit()

    cilo_all = Cilo.query.all()
    graph_array = []
    graph_array.append({"id": "root", "isroot": True, "topic": "<b>%s</b>" % (cilo.name)}, )
    if cilo.pre_cilo1:
        graph_array.append({"id": "sub1", "parentid": "root", "topic": Cilo.query.get(cilo.pre_cilo1).name}, )
    if cilo.pre_cilo2:
        graph_array.append({"id": "sub2", "parentid": "root", "topic": Cilo.query.get(cilo.pre_cilo2).name}, )
    if cilo.pre_cilo3:
        graph_array.append({"id": "sub3", "parentid": "root", "topic": Cilo.query.get(cilo.pre_cilo3).name}, )

    graph_json = json.dumps(graph_array)
    print(graph_json)
    return render_template('edit_cilo.html',
                           title=errmsg,
                           cilo=cilo,
                           cilo_all=cilo_all,
                           graph_json=graph_json,
                           )


def parse_excel(stream):
    wb = load_workbook(io.BytesIO(stream.read()))
    rows = list(wb[wb.sheetnames[0]].rows)
    title = rows[0]
    data = rows[1:]
    grade_list = []
    for _data in data:
        _ = dict()
        for k, v in zip(title, _data):
            if k.value is None or v.value is None:
                continue
            _[str(k.value).strip()] = str(v.value).strip()
        if len(_) == 0:
            continue
        grade_list.append(_)
    return grade_list


@coursedesignerBP.route('visualize_dep', methods=['GET'])
def visualize_dep():
    keyword = request.args.get("programme")
    return visualize_inner(keyword)
